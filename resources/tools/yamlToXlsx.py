#!/usr/bin/env python3
import sys
from argparse import ArgumentParser, Namespace
from datetime import datetime
import yaml
import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.builtins import percent

# Find data to add (Ansible and IBD)
#   Execution Date
#   Build number
#   Cumulative Number of executions
#   CCSM drop
#   CCSM version
#   Total Number of TCs
#      Number of executed TCs
#      Successful TCs
#      Cluster-related Failures
#      Failed TCs (unstable)
#      Failed TCs (CCSM)
#      Failed TCs (ADP)
#      Uninstallation TS Failures
#      Installation TS Failures
#      Traffic Mix TS Failures
#      Stability TS Failures
#      Maintainability TS Failures
#      Robustness TS Failures
#      CommonFailures TS Failures
#      PodsFaults TS Failures
#      Licenses TS Failures
#      Overload TS Failures



HEADER_ROW = 20
START_COLUMN = {
    'ansible': {
        'mTLS': {
            'IPv4': 'A',
            'IPv6': None
        },
        'CLEAR': {
            'IPv4': 'AW',
            'IPv6': None
        }
    },
    'IBD': {
        'mTLS': {
            'IPv4': 'X',
            'IPv6': 'CT'
        },
        'CLEAR': {
            'IPv4': 'BU',
            'IPv6': None
        }
    }
}

GREEN = "00b050"
BLUE  = "1f4e79"
WHITE = "FFFFFF"


def parse_arguments():
    '''Read and parse paramaters received in the command line'''
    parser = ArgumentParser(description=__doc__)
    parser.add_argument('-i', '--input', default='summary_results.yaml', help='Yaml file including JCAT TC results')
    parser.add_argument('-n', '--filename', default="NFT_Automation_TCs.xlsx", help='Confluence attachment XLSX file name to update')
    parser.add_argument('-o', '--pageid', default="814408112", help='Confluence page ID where files are attached')
    parser.add_argument('--clear', action='store_true', help='Clear HTTP')
    parser.add_argument('--ipv6', action='store_true', help='IPv6')
    parser.add_argument('-u', '--userid', default='esdccci', help='UserID to access target URL Rest API')
    parser.add_argument('-p', '--password', default='', help='Password to access target URL Rest API')
    parser.add_argument('--dry-run', action='store_true', help='Do not post result to Confluence')
    args = parser.parse_args()
    return args


def download_workbook(parms: Namespace) -> Workbook:
    '''Get a file from an URL using its Rest API'''
    workbook = None
    url = "https://eteamspace.internal.ericsson.com/download/attachments"
    response = requests.get(f"{url}/{parms.pageid}/{parms.filename}?api=v2", auth=HTTPBasicAuth(parms.userid, parms.password))
    if (response.status_code == 200):
        with open(parms.filename, 'wb') as fp:
            fp.write(response.content)

        workbook = openpyxl.load_workbook(parms.filename)
    return workbook


def upload_workbook(parms: Namespace) -> int:
    '''Upload updated XLSX file to Confluence'''
    url = f"https://eteamspace.internal.ericsson.com/pages/doattachfile.action?pageId={parms.pageid}"
    headers = {'X-Atlassian-Token': 'no-check'}
    files = {parms.filename:(parms.filename, open(parms.filename, 'rb'))}
    res = requests.post(url, headers=headers, auth=HTTPBasicAuth(parms.userid, parms.password), files=files)
    print(f"[INFO] Upload result: {res.status_code}")
    return 0


def get_columns_to_update(data_item: dict, clear: bool, ipv6: bool) -> list:
    '''Find column range to update'''
    cluster = 'ansible' if '-ans-' in data_item['environment'] else 'IBD'
    security = 'CLEAR' if clear else 'mTLS'
    ipver = 'IPv6' if ipv6 else 'IPv4'

    cols = [START_COLUMN[cluster][security][ipver]]

    # Column start letter + number of keys in input data + 1 (autocalc 'Cumulative number of executions' column)
    cols.append(get_column_letter(column_index_from_string(cols[0]) + len(data_item.keys())))
    return cols


def get_row_to_update(sheet: Worksheet, colinds: list) -> int:
    '''Find row index to update'''
    rowindex = HEADER_ROW+1
    for row in sheet.iter_rows(HEADER_ROW+1, sheet.max_row+1, column_index_from_string(colinds[0]), column_index_from_string(colinds[1])):
        if row[0].value in (None, ''):
            rowindex = row[0].row
            break
    return rowindex


def total_stable_percent_cell(sheet: Worksheet, data_item: dict, colinds: list, rowind: int) -> None:
    '''Set format and value for "Stability CL2 (stable/Total)" cell'''
    totNum = data_item['Total Number of TCs']
    failTcs = data_item['Failed TCs (unstable)']
    cellIndex = f"{get_column_letter(column_index_from_string(colinds[1]))}{rowind}"
    sheet[cellIndex].value = f"=1-({failTcs}/{totNum})"
    sheet[cellIndex].number_format = '0.0%'
    sheet[cellIndex].fill = PatternFill("solid", fgColor=GREEN)
    sheet[cellIndex].font = Font(name='Calibri', size=11, color=WHITE)


def average_stable_percent_cell(sheet: Worksheet, colinds: list, rowind: int) -> None:
    '''Set format and value for "Stability CL2 Rolling average (last 10 executions)" cell'''
    avgFirst = f"{get_column_letter(column_index_from_string(colinds[1]))}{rowind-10}"
    avgLast  = f"{get_column_letter(column_index_from_string(colinds[1]))}{rowind}"
    cellIndex = f"{get_column_letter(column_index_from_string(colinds[1])+1)}{rowind}"
    sheet[cellIndex].value = f"=AVERAGE({avgFirst},{avgLast})"
    sheet[cellIndex].number_format = '0.0%'
    sheet[cellIndex].fill = PatternFill("solid", fgColor=BLUE)
    sheet[cellIndex].font = Font(name='Calibri', size=11, color="FFFFFF")


############
#   MAIN   #
############
if __name__ == '__main__':
    result = 0
    params = parse_arguments()
    print("[INFO] Reading Jenkins build data")
    data = yaml.safe_load(open(params.input))
    cancel = False

    print(f"[INFO] Opening workbook: '{params.filename}'")
    workbook = download_workbook(params)

    for item in data:
        print(f"[INFO] Processing spreadsheet: '{item}'")
        sheet = workbook.get_sheet_by_name(item)

        print("[INFO] Locating cell indexes to update")
        colindexes = get_columns_to_update(data[item], params.clear, params.ipv6)
        rowindex = get_row_to_update(sheet, colindexes)

        print("[INFO] Updating cells")
        for col in sheet.iter_cols(min_row=HEADER_ROW, max_row=HEADER_ROW,
                                   min_col=column_index_from_string(colindexes[0]),
                                   max_col=column_index_from_string(colindexes[1])):
            for cell in col:
                if cell.value == "Build number":
                    if sheet[f"{get_column_letter(cell.column)}{rowindex-1}"] == data[item][cell.value]:
                        print("[WARNING] Data seems to be already included. Aborting.")
                        cancel = True
                        break
                if cell.value == "Cumulative Number of executions":
                    prev_row = rowindex-1 if rowindex > HEADER_ROW+1 else HEADER_ROW+1
                    sheet[f"{get_column_letter(cell.column)}{rowindex}"] = int(sheet[f"{get_column_letter(cell.column)}{prev_row}"].value)+1
                elif cell.value in data[item]:
                    value = data[item][cell.value] if data[item][cell.value] != "" else "-"
                    if 'date' in cell.value.lower():
                        value = datetime.strptime(datetime.strftime(value, '%Y/%d/%m %H:%M:%S'), '%Y/%d/%m %H:%M:%S')
                        sheet[f"{get_column_letter(cell.column)}{rowindex}"].number_format = 'd-mmm'
                    else:
                        sheet[f"{get_column_letter(cell.column)}{rowindex}"].alignment = Alignment(horizontal="center")
                    sheet[f"{get_column_letter(cell.column)}{rowindex}"] = value
            if cancel:
                break
        else:
            total_stable_percent_cell(sheet, data[item], colindexes, rowindex)
            average_stable_percent_cell(sheet, colindexes, rowindex)

    if not cancel:
        print("[INFO] Saving changes to file")
        workbook.save(params.filename)

        if not params.dry_run:
            print("[INFO] Upload file to Confluence")
            result = upload_workbook(params)

    exit(result)
